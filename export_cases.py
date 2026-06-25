#!/usr/bin/env python3
"""
export_cases.py — Generate a formatted Excel table from extracted art/design history cases.

Usage:
    python export_cases.py --input cases_all_v2.json --output cases.xlsx
    python export_cases.py --data '[{...}]' --output cases.xlsx

Current JSON schema (cases_all_v2.json):
[
  {
    "类别": "国内/国外",
    "年代": "...",
    "排序年": 1234,
    "作品/事件名称": "...",
    "作者/设计师": "...",
    "来源书目": "《书名》",
    "简介": "...",
    "原文引用": "...",
    "教学分析": "...",
    "批注": "...",
    "标签": ["1.2 二元与多元"],
    "importance": 0,
    "image": "",
    "link": ""
  }
]
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


COLUMNS = [
    ("作品/事件名称", 30),
    ("年代",         12),
    ("类别",          8),
    ("作者/设计师",   22),
    ("标签",          28),
    ("来源书目",      20),
    ("原文引用",      45),
    ("教学分析",      55),
    ("批注",          30),
    ("重要性",        10),
]

# Map JSON keys to column headers
KEY_MAP = {
    "作品/事件名称": "作品/事件名称",
    "年代": "年代",
    "类别": "类别",
    "作者/设计师": "作者/设计师",
    "标签": "标签",          # array → joined string
    "来源书目": "来源书目",
    "原文引用": "原文引用",
    "教学分析": "教学分析",
    "批注": "批注",
    "importance": "重要性",
}

HEADER_FILL   = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ALT_FILL      = PatternFill("solid", fgColor="F2F4F6")
AMBER_FILL    = PatternFill("solid", fgColor="FFF3CD")
WRAP          = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP   = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN          = Side(style="thin", color="CCCCCC")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

import re as _re
def _strip_bold(text: str) -> str:
    return _re.sub(r'\*\*(.+?)\*\*', r'\1', text)


def load_cases(data_str, input_file):
    if data_str:
        return json.loads(data_str)
    if input_file:
        return json.loads(Path(input_file).read_text(encoding="utf-8"))
    raise ValueError("Provide --data or --input")


def normalize(case: dict) -> dict:
    """Remap JSON keys to canonical column headers; handle arrays and bold markers."""
    out = {}
    for k, v in case.items():
        col = KEY_MAP.get(k)
        if col is None:
            continue
        if isinstance(v, list):
            # 标签 array → "1.2 二元与多元 / 4.1 形式与功能"（空数组→空字符串）
            out[col] = " / ".join(str(i) for i in v) if v else ""
        elif v is None:
            out[col] = ""
        else:
            text = str(v)
            if col in ("教学分析", "原文引用"):
                text = _strip_bold(text)
            out[col] = text
    return out


def build_workbook(cases: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "案例库"

    headers = [col[0] for col in COLUMNS]
    widths  = [col[1] for col in COLUMNS]

    # Header row
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_WRAP
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, raw_case in enumerate(cases, start=2):
        case = normalize(raw_case)
        fill = ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, header in enumerate(headers, start=1):
            value = case.get(header, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
            cell.border    = BORDER
            # 批注列：非空时琥珀色底
            if header == "批注" and value:
                cell.fill = AMBER_FILL
            elif fill:
                cell.fill = fill

        # Auto row height hint (openpyxl can't truly auto-fit, set a generous min)
        max_lines = max(
            len(str(case.get(h, "")).split("\n")) for h in headers
        )
        ws.row_dimensions[row_idx].height = max(18, min(max_lines * 15, 120))

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    return wb


def main():
    parser = argparse.ArgumentParser(description="Export art history cases to Excel")
    parser.add_argument("--data",   help="JSON array string of cases")
    parser.add_argument("--input",  help="Path to JSON file of cases")
    parser.add_argument("--output", default="cases.xlsx", help="Output .xlsx path")
    args = parser.parse_args()

    cases = load_cases(args.data, args.input)
    wb    = build_workbook(cases)
    wb.save(args.output)
    print(f"Saved {len(cases)} cases → {args.output}")


if __name__ == "__main__":
    main()
